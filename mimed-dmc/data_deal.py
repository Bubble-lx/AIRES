import os
from openpyxl import load_workbook

def find_xlsx_files(path):
    # Search for all xlsx files in the specified directory
    xlsx_files = [os.path.join(path, f) for f in os.listdir(path) if f.endswith('.xlsx')]
    return xlsx_files

def get_sheets_from_xlsx(file_path):
    # Load Excel file
    workbook = load_workbook(filename=file_path, data_only=True)
    # Get the names of all sheets
    sheets = workbook.sheetnames
    return sheets

def main():
    path = "/deir-main/src/logs/Results"

    # Retrieve all Excel files in the directory
    xlsx_files = find_xlsx_files(path)

    # Output sheets for each file
    for file_path in xlsx_files:
        sheets = get_sheets_from_xlsx(file_path)
        print(f"File: {file_path} has the following sheets: {sheets}")


if __name__ == "__main__":
    main()
